import csv
from datetime import datetime
import re
from textwrap import fill
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import pdfkit
from jinja2 import Environment, FileSystemLoader

#класс Зарплата
class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
    # получение зп в рублях
    def get_salary_in_rub(self):
        return (float(self.salary_from) + float(self.salary_to)) / 2 * currency_to_rub[self.salary_currency]

#класс вакансия
class Vacancy:
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

#класс обработки входных данных
class InputData:
    # ввод параметров
    @staticmethod
    def input_parameters():
        file_name = input('Введите название файла: ')
        vacancy_name = input('Введите название профессии: ')
        return file_name, vacancy_name

    def make_data(self):
        input_parameters = InputData.input_parameters()
        if input_parameters is not None:
            file_name, vacancy_name = input_parameters
            vacancies_objects = DataSet(file_name).vacancies_objects
            InputData.print_transform_data(vacancies_objects, vacancy_name)

    #получение словаря зарплат
    @staticmethod
    def get_salary_dictionary(dictionary):
        for key, value in dictionary.items():
            if len(value) == 0:
                dictionary[key] = 0
            else:
                dictionary[key] = int(sum(value) / len(value))
        return dictionary

    #вывод преобразованных данных
    @staticmethod
    def print_transform_data(vacancies_objects, vacancy_name):
        vacancies_dictionary = vacancies_objects
        years = set()
        for vacancy in vacancies_dictionary:
            years.add(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y")))
        area_dictionary, years_count_dictionary, years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict = InputData.get_new_parameters(
            years)
        InputData.get_correct_vacancies(area_dictionary, vacancies_dictionary, vacancy_name, years_count_dictionary,
                                        years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict)

        years_salary_dictionary = InputData.get_salary_dictionary(years_salary_dictionary)
        years_salary_vacancy_dict = InputData.get_salary_dictionary(years_salary_vacancy_dict)
        area_list = area_dictionary.items()
        area_list = [area for area in area_list if len(area[1]) >= len(vacancies_dictionary) // 100]
        area_salary_dict = sorted(area_list, key=lambda area: sum(area[1]) / len(area[1]), reverse=True)
        area_count_dict = sorted(area_list, key=lambda item: len(item[1]) / len(vacancies_dictionary), reverse=True)
        area_salary_dict = {item[0]: int(sum(item[1]) / len(item[1]))
                            for item in area_salary_dict[0: min(len(area_salary_dict), 10)]}
        area_count_dict = {item[0]: round(len(item[1]) / len(vacancies_dictionary), 4)
                           for item in area_count_dict[0: min(len(area_count_dict), 10)]}

        InputData.print_data(area_count_dict, area_salary_dict, years_count_dictionary, years_count_vacancy_dict,
                             years_salary_dictionary, years_salary_vacancy_dict)
        analytic_year.append(years_salary_dictionary)
        analytic_year.append(years_salary_vacancy_dict)
        analytic_year.append(years_count_dictionary)
        analytic_year.append(years_count_vacancy_dict)
        analytic_city .append(area_salary_dict)
        analytic_city_new.append(area_count_dict)

    @staticmethod
    def get_new_parameters(years):
        years = list(range(min(years), max(years) + 1))
        years_salary_dictionary = {year: [] for year in years}
        years_salary_vacancy_dict = {year: [] for year in years}
        years_count_dictionary = {year: 0 for year in years}
        years_count_vacancy_dict = {year: 0 for year in years}
        area_dictionary = {}
        return area_dictionary, years_count_dictionary, years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict

    #получение корректных вакансий
    @staticmethod
    def get_correct_vacancies(area_dict, vacancies_dictionary, vacancy_name, years_count_dictionary,
                              years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict):
        for vacancy in vacancies_dictionary:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y"))
            years_salary_dictionary[year].append(vacancy.salary.get_salary_in_rub())
            years_count_dictionary[year] += 1
            if vacancy_name in vacancy.name:
                years_salary_vacancy_dict[year].append(vacancy.salary.get_salary_in_rub())
                years_count_vacancy_dict[year] += 1
            if vacancy.area_name in area_dict:
                area_dict[vacancy.area_name].append(vacancy.salary.get_salary_in_rub())
            else:
                area_dict[vacancy.area_name] = [vacancy.salary.get_salary_in_rub()]

    @staticmethod
    def print_data(area_count_dict, area_salary_dict, years_count_dictionary, years_count_vacancy_dict,
                   years_salary_dictionary, years_salary_vacancy_dict):
        print(f'Динамика уровня зарплат по годам: {years_salary_dictionary}')
        print(f'Динамика количества вакансий по годам: {years_count_dictionary}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {years_salary_vacancy_dict}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {years_count_vacancy_dict}')
        print(f'Уровень зарплат по городам (в порядке убывания): {area_salary_dict}')
        print(f'Доля вакансий по городам (в порядке убывания): {area_count_dict}')


#класс работы с данными
class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.parser_csv(file_name)
    #чтение файла
    @staticmethod
    def csv_reader(file_name):
        with open(file_name, encoding='utf_8_sig') as file:
            reader = [row for row in csv.reader(file)]
            try:
                name = reader.pop(0)
                return name, reader
            except:
                print('Пустой файл')
                exit()
    #удаление символов из строки
    @staticmethod
    def clear_string(str_value):
        return ' '.join(re.sub(r"<[^>]+>", '', str_value).split())
    #парсер
    @staticmethod
    def parser_csv(file_name):
        naming, reader = DataSet.csv_reader(file_name)
        dic_vacancies = []
        filtered_vacancies = [x for x in reader if len(x) == len(naming) and '' not in x]
        for row in filtered_vacancies:
            dictionary = {}
            for i in range(0, len(row)):
                if row[i].find("\n") != -1:
                    answer = [DataSet.clear_string(element) for element in row[i].split('\n')]
                else:
                    answer = DataSet.clear_string(row[i])
                dictionary[naming[i]] = answer
            dic_vacancies.append(
                Vacancy(dictionary['name'],
                        Salary(dictionary['salary_from'],
                               dictionary['salary_to'],
                               dictionary['salary_currency']),
                               dictionary['area_name'],
                               dictionary['published_at']))
        return dic_vacancies

#класс для формирования ответа программы в файле excel
class Report:
    def __init__(self, sheet_title_year, sheet_title_city, color_border, style_border, bold_text):
        self.sheet_title_year = sheet_title_year
        self.sheet_title_city = sheet_title_city
        self.color_border = color_border
        self.style_border = style_border
        self.bold_text = bold_text

    #добавление значений
    @staticmethod
    def append_new_values(list_dict, work, style_border, num_column = 1):
        for i, element in enumerate(list_dict[0].keys()):
            work.cell(row = i + 2, column = num_column).value = element
            work.cell(row = i + 2, column = num_column).border = style_border
        number_col = num_column + 1
        for dictionary in list_dict:
            for i, element in enumerate(dictionary.values()):
                if type(element) == float:
                    work.cell(row = i + 2, column = number_col).number_format = '0.00%'
                work.cell(row = i + 2, column = number_col).value = element
                work.cell(row = i + 2, column = number_col).border = style_border
            number_col += 1

    @staticmethod
    def get_width_column(sheet, font_size=11):
        columns_dictionary = {}
        for row in sheet.rows:
            for cell in row:
                letter = cell.column_letter
                if cell.value:
                    len_cell = len(str(cell.value))
                    cell_dict = 0
                    if letter in columns_dictionary:
                        cell_dict = columns_dictionary[letter]
                    if len_cell > cell_dict:
                        columns_dictionary[letter] = len_cell
                        new_width_col = len_cell * font_size ** (font_size * 0.009)
                        sheet.column_dimensions[cell.column_letter].width = new_width_col

    def prapare_excel_data(self, field_statistic_city, field_statistic_year, style_border):
        bold_heading = Font(bold=self.bold_text)
        border = Border(top = style_border, bottom = style_border, left = style_border, right = style_border)
        work_book = Workbook()
        work_year = work_book.active
        work_year.title = self.sheet_title_year
        work_city = work_book.create_sheet(self.sheet_title_city)
        work_year.append(field_statistic_year)
        work_city.append(field_statistic_city)
        Report.append_new_values(analytic_year, work_year, border)
        Report.append_new_values(analytic_city, work_city, border)
        Report.append_new_values(analytic_city_new, work_city, border, 3)
        return bold_heading, border, work_book, work_city, work_year

    def generate_excel_file(self, vacancy_name):
        field_year = ['Год',
                      'Средняя зарплата',
                      f'Средняя зарплата - {vacancy_name}',
                      'Количество вакансий',
                      f'Количество вакансий - {vacancy_name}']
        field_city = ['Город',
                      'Уровень зарплат',
                      'Город',
                      'Доля вакансий']
        style = Side(border_style = self.style_border, color = self.color_border)
        bold_heading, border, work_book, work_sheet_city, work_sheet_year = self.prapare_excel_data(
            field_city, field_year, style)
        for x in work_book:
            for cell in x[1]:
                cell.font = bold_heading
                cell.border = border
        work_sheet_city.insert_cols(3, 1)
        Report.get_width_column(work_sheet_year)
        Report.get_width_column(work_sheet_city)
        work_book.save('report.xlsx')

    #создание круговой диаграммы
    @staticmethod
    def generate_pie_chart(dictionary, title, ax):
        label = ['Другие']
        for city in dictionary.keys():
            label.append(city)
        sizes_list = [1 - sum(dictionary.values())]
        for size in dictionary.values():
            sizes_list.append(size)
        textprops = {"fontsize": 6}
        ax.set_title(title)
        ax.pie(sizes_list, labels=label, textprops=textprops)
        ax.axis('equal')

    #создание горизонтальной столбчатой диаграммы
    @staticmethod
    def generate_horizontal_bar_chart(dictionary, title, ax):
        plt.rcdefaults()
        cities = []
        for city in dictionary.keys():
            if ' ' in city:
                cities.append(city.replace(' ', '\n'))
            elif '-' in city:
                cities.append(city.replace('-', '-\n'))
            else:
                cities.append(city)
        Report.prepare_horizontal_bar(ax, cities, dictionary, title)

   # создание вертикальной столбчатой диаграммы
    @staticmethod
    def generate_group_bar_chart(first_dict, second_dict, first_label, second_label, title, vacancy_name, ax):
        new_vacancy = first_dict.values()
        current_vacancy = second_dict.values()
        label = first_dict.keys()
        x = np.arange(len(label))
        width = 0.35
        ax.bar(x - width / 2, new_vacancy, width, label=first_label)
        ax.bar(x + width / 2, current_vacancy, width, label=fill(f'{second_label} {vacancy_name.lower()}', 20))
        ax.set_title(title)
        ax.set_xticks(x, label, rotation=90)
        ax.legend(fontsize=8, loc='upper left')
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(visible=True, axis='y')

    @staticmethod
    def prepare_horizontal_bar(ax, cities, dictionary, title):
        y_pos = np.arange(len(cities))
        performance = dictionary.values()
        ax.barh(y_pos, performance)
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.set_title(title)
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.grid(visible=True, axis='x')

    @staticmethod
    def generate_image(vacancy_name):
        fig, ax = plt.subplots(2, 2)
        Report.generate_group_bar_chart(analytic_year[0],
                                        analytic_year[1],
                                        'средняя з/п',
                                        'з/п',
                                        'Уровень зарплат по годам',
                                        vacancy_name,
                                        ax[0][0])
        Report.generate_group_bar_chart(analytic_year[2],
                                        analytic_year[3],
                                        'Количество вакансий',
                                        'Количество вакансий',
                                        'Количество вакансий по годам',
                                        vacancy_name,
                                        ax[0][1])
        Report.generate_horizontal_bar_chart(analytic_city[0], 'Уровень зарплат по городам', ax[1][0])
        Report.generate_pie_chart(analytic_city_new[0], 'Доля вакансий по городам', ax[1][1])
        plt.tight_layout()
        plt.savefig('graph.png')
        plt.show()

    @staticmethod
    def generate_pdf(vacancy_name):
        table1 = 'Статистика по годам'
        table2 = 'Статистика по городам'
        field1 = ['Город', 'Уровень зарплат']
        field2 = ['Город', 'Доля вакансий']

        field_statistic_year = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancy_name}',
                                'Количество вакансий', f'Количество вакансий - {vacancy_name}']
        years = analytic_year[0].keys()
        image_url = 'C:\\Users\\RBT\\PycharmProjects\\Korelina\\graph.png'
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")
        dict_city_part = {}
        for key, value in analytic_city_new[0].items():
            value = f'{round(value * 100, 2)}%'
            dict_city_part[key] = value
        pdf_template = template.render({'vacancy_name': vacancy_name, 'image_url': image_url,
                                        'field_statistic_year': field_statistic_year, 'years': years,
                                        'analytic_year': analytic_year,
                                        'table1': table1, 'table2': table2,
                                        'field1': field1,
                                        'field2': field2,
                                        'analytic_city': analytic_city[0],
                                        'dict_city_part': dict_city_part})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Users\RBT\Downloads\wkhtmltox\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


words_dictionary = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary_from': 'Нижняя граница вилки оклада',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary_gross': 'Оклад указан до вычета налогов',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии',
    'salary': 'Оклад'}

work_experience = {
    'noExperience': 'Нет опыта',
    'between1And3': 'От 1 года до 3 лет',
    'between3And6': 'От 3 до 6 лет',
    'moreThan6': 'Более 6 лет'}

new_currency = {'AZN': 'Манаты',
                'BYR': 'Белорусские рубли',
                'EUR': 'Евро',
                'GEL': 'Грузинский лари',
                'KGS': 'Киргизский сом',
                'KZT': 'Тенге', 'RUR': 'Рубли',
                'UAH': 'Гривны',
                'USD': 'Доллары',
                'UZS': 'Узбекский сум'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

experience_order = {
    'noExperience': 0,
    'between1And3': 1,
    'between3And6': 2,
    'moreThan6': 3,
}

bool_to_str = {
    'False': 'Нет',
    'True': 'Да'
}

analytic_year = []
analytic_city = []
analytic_city_new = []

def main():
    a = InputData()
    a.make_data()

if __name__ == '__main__':
    main()





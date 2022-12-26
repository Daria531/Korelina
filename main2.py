import csv
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

#класс Зарплата
class Salary:
    """
    Класс для предоставления зарплаты.

    Attributes:
       salary_from (int): Нижняя граница вилки оклада
       salary_to (int): Верхняя граница вилки оклада
       salary_currency (str): Идентификатор валюты оклада
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """
        Инициализирует объект Salary, выполняет конвертацию для целочисленных полей.

        Args:
            salary_from (str or int or float): Нижняя граница вилки оклада
            salary_to (str or int or float): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
    # получение зп в рублях
    def get_salary_in_rub(self):
        """
        Вычисляет среднюю зарплату из вилки и переводит в рубли при помощи словаря - currency_to_rub.
        Returns:
            float: Средняя зарплата в рублях
        """
        return (float(self.salary_from) + float(self.salary_to)) / 2 * currency_to_rub[self.salary_currency]

#класс вакансия
class Vacancy:
    """
    Класс для представления вакансии.

    Attributes:
        name (str): Название вакансии
        salary (int): Средняя зарплата
        area_name (str): Название региона
        published_at (str): Дата публикации вакансии
    """
    def __init__(self, name, salary, area_name, published_at):
        """
        Инициализирует объект Vacancy

        Args:
            name (str): Название вакансии
            salary (str or int or float): Средняя зарплата
            area_name (str): Название региона
            published_at (str): Дата публикации вакансии
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

#класс обработки входных данных
class InputData:
    """
    Обработка вводимых параметров, печать статистки, создание графиков.
    """
    # ввод параметров
    @staticmethod
    def input_parameters():
        """
        Получает название файла и название вакансии.
        Returns:
            str: CSV файл
            str: Название вакансии
        """
        file_name = input('Введите название файла: ')
        vacancy_name = input('Введите название профессии: ')
        return file_name, vacancy_name

    def make_data(self):
        """Печать статистики на экран, создание таблиц, графиков"""
        input_parameters = InputData.input_parameters()
        if input_parameters is not None:
            file_name, vacancy_name = input_parameters
            vacancies_objects = DataSet(file_name).vacancies_objects
            InputData.print_transform_data(vacancies_objects, vacancy_name)

    #получение словаря зарплат
    @staticmethod
    def get_salary_dictionary(dictionary):
        """
        Преобразование словаря (ключ - год, значение - средняя зарплата)
        Args:
            dictionary (list): Словарь (ключ - год, значение - средняя зарплата)

        Returns:
            dict (list): Средняя зарплата за год (ключ - год, значение - средняя зарплата)
        """
        for key, value in dictionary.items():
            if len(value) == 0:
                dictionary[key] = 0
            else:
                dictionary[key] = int(sum(value) / len(value))
        return dictionary

    #вывод преобразованных данных
    @staticmethod
    def print_transform_data(vacancies_objects, vacancy_name):
        """
        Печатать статистики зарплаты и вакансий
        Args:
            vacancies_objects (list): Список с вакансиями
            vacancy_name (str): Название вакансии для статистики
        """
        vacancies_dictionary = vacancies_objects
        years = set()
        for vacancy in vacancies_dictionary:
            years.add(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y")))
        area_dictionary, years_count_dictionary, years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict = InputData.get_new_parameters(
            years)
        InputData.get_correct_vacancies(area_dictionary, vacancies_dictionary, vacancy_name, years_count_dictionary,
                                        years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict)

        years_salary_dictionary = InputData.get_salary_dictionary(years_salary_dictionary)
        years_salary_vacancy_dict = InputData.get_salary_dictionary(years_salary_vacancy_dict)

        area_list = area_dictionary.items()
        area_list = [area for area in area_list if len(area[1]) >= len(vacancies_dictionary) // 100]
        area_salary_dict = sorted(area_list, key=lambda area: sum(area[1]) / len(area[1]), reverse=True)
        area_count_dict = sorted(area_list, key=lambda item: len(item[1]) / len(vacancies_dictionary), reverse=True)
        area_salary_dict = {item[0]: int(sum(item[1]) / len(item[1]))
                            for item in area_salary_dict[0: min(len(area_salary_dict), 10)]}
        area_count_dict = {item[0]: round(len(item[1]) / len(vacancies_dictionary), 4)
                           for item in area_count_dict[0: min(len(area_count_dict), 10)]}

        InputData.print_data(area_count_dict, area_salary_dict, years_count_dictionary, years_count_vacancy_dict,
                             years_salary_dictionary, years_salary_vacancy_dict)

        list_transform_dictionary.append(years_salary_dictionary)
        list_transform_dictionary.append(years_salary_vacancy_dict)
        list_transform_dictionary.append(years_count_dictionary)
        list_transform_dictionary.append(years_count_vacancy_dict)
        list_transform_city.append(area_salary_dict)
        list_transform_city_new.append(area_count_dict)

    @staticmethod
    def get_new_parameters(years):
        """
        Вспомогательный метод создания параметров для печати статистики.
        Args:
            years: Список лет, по которым будет проведена статистика
        """
        years = list(range(min(years), max(years) + 1))
        years_salary_dictionary = {year: [] for year in years}
        years_salary_vacancy_dict = {year: [] for year in years}
        years_count_dictionary = {year: 0 for year in years}
        years_count_vacancy_dict = {year: 0 for year in years}
        area_dictionary = {}
        return area_dictionary, years_count_dictionary, years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict

    #получение корректных вакансий
    @staticmethod
    def get_correct_vacancies(area_dict, vacancies_dictionary, vacancy_name, years_count_dictionary,
                              years_count_vacancy_dict, years_salary_dictionary, years_salary_vacancy_dict):
        """
        Получение корректных данных вакансий
        Args:
            area_dict (list): Список вакансий в границах
            vacancies_dictionary (list): Список вакансий
            vacancy_name (str): Название вакансии
            years_count_dictionary (list): Динамика количества вакансий по годам
            years_count_vacancy_dict (list): Динамика количества вакансий по годам для выбранной профессии
            years_salary_dictionary (list): Динамика уровня зарплат по годам
            years_salary_vacancy_dict (list): Динамика уровня зарплат по годам для выбранной профессии
        """
        for vacancy in vacancies_dictionary:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime("%Y"))
            years_salary_dictionary[year].append(vacancy.salary.get_salary_in_rub())
            years_count_dictionary[year] += 1
            if vacancy_name in vacancy.name:
                years_salary_vacancy_dict[year].append(vacancy.salary.get_salary_in_rub())
                years_count_vacancy_dict[year] += 1
            if vacancy.area_name in area_dict:
                area_dict[vacancy.area_name].append(vacancy.salary.get_salary_in_rub())
            else:
                area_dict[vacancy.area_name] = [vacancy.salary.get_salary_in_rub()]

    @staticmethod
    def print_data(area_count_dict, area_salary_dict, years_count_dictionary, years_count_vacancy_dict,
                   years_salary_dictionary, years_salary_vacancy_dict):
        """
        Печать данных о вакансиях.
        Args:
            area_count_dict (list): Список вакансий в границах
            area_salary_dict (list): Список зарплат в границах
            years_count_dictionary (list): Динамика количества вакансий по годам
            years_count_vacancy_dict (list): Динамика количества вакансий по годам для выбранной профессии
            years_salary_dictionary (list): Динамика уровня зарплат по годам
            years_salary_vacancy_dict (list): Динамика уровня зарплат по годам для выбранной профессии
        """
        print(f'Динамика уровня зарплат по годам: {years_salary_dictionary}')
        print(f'Динамика количества вакансий по годам: {years_count_dictionary}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {years_salary_vacancy_dict}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {years_count_vacancy_dict}')
        print(f'Уровень зарплат по городам (в порядке убывания): {area_salary_dict}')
        print(f'Доля вакансий по городам (в порядке убывания): {area_count_dict}')


#класс работы с данными
class DataSet:
    """
    Класс для получения данных csv-файла.
    Attributes:
        file_name (str): Название csv-файла
        vacancies_objects (list): Список с вакансиями
    """
    def __init__(self, file_name):
        """
        Инициализирует объект DataSet, получает список с вакансиями.
        Args:
            file_name: Название csv-файла
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.parser_csv(file_name)
    #чтение файла
    @staticmethod
    def csv_reader(file_name):
        """
        Чтение csv-файла, создание списка вакансий.
        Args:
            file_name (str): Название csv-файла

        Returns:
            list: Список вакансий
        """
        with open(file_name, encoding='utf_8_sig') as file:
            reader = [row for row in csv.reader(file)]
            try:
                name = reader.pop(0)
                return name, reader
            except:
                print('Пустой файл')
                exit()
    #удаление символов из строки
    @staticmethod
    def clear_string(str_value):
        """
        Удаление тегов из строки
        Args:
            str_value (str): Строка

        Returns:
            str (str): Очищенная от тегов строка
        """
        return ' '.join(re.sub(r"<[^>]+>", '', str_value).split())
    #парсер
    @staticmethod
    def parser_csv(file_name):
        """
        Парсинг данных csv-файла
        Args:
            file_name (str): Название csv-файла

        Returns:
            dic_vacancies (list): Список обработанных данных
        """
        naming, reader = DataSet.csv_reader(file_name)
        dic_vacancies = []
        filtered_vacancies = [x for x in reader if len(x) == len(naming) and '' not in x]
        for row in filtered_vacancies:
            dictionary = {}
            for i in range(0, len(row)):
                if row[i].find("\n") != -1:
                    answer = [DataSet.clear_string(element) for element in row[i].split('\n')]
                else:
                    answer = DataSet.clear_string(row[i])
                dictionary[naming[i]] = answer
            dic_vacancies.append(
                Vacancy(dictionary['name'],
                        Salary(dictionary['salary_from'],
                               dictionary['salary_to'],
                               dictionary['salary_currency']),
                               dictionary['area_name'],
                               dictionary['published_at']))
        return dic_vacancies

#класс для формирования ответа программы в файле excel
class Report:
    """
    Класс для создания таблиц, графиков и отчета по статистике.
    Attributes:
        sheet_title_year (str): Список с таблицей статистики по годам
        sheet_title_city (str): Список с таблицей статистики по городам
        color_border (str): Цвет обводки ячеек таблицы
        style_border (str): Толщина обводки ячеек таблицы
        bold_text (bool): Жирность текста
    """
    def __init__(self, sheet_title_year, sheet_title_city, color_border, style_border, bold_text):
        """
        Инициализация объекта класса Report.
        Args:
            sheet_title_year (str): Список с таблицей статистики по годам
            sheet_title_city (str): Список с таблицей статистики по городам
            color_border (str): Цвет обводки ячеек таблицы
            style_border (str): Толщина обводки ячеек таблицы
            bold_text (bool): Жирность текста
        """
        self.sheet_title_year = sheet_title_year
        self.sheet_title_city = sheet_title_city
        self.color_border = color_border
        self.style_border = style_border
        self.bold_text = bold_text
    #добавление значений
    @staticmethod
    def append_values(list_dict, work, style_border, num_column = 1):
        """
        Добавление значений (вспомогательный метод).
        Args:
            list_dict (list): Список значений
            work (list): Рабочая область
            style_border (str): Толщина обводки ячеек таблицы
            num_column (int): Номер колонки (по умолчанию 1)
        """
        for i, element in enumerate(list_dict[0].keys()):
            work.cell(row =i + 2, column = num_column).value = element
            work.cell(row =i + 2, column = num_column).border = style_border
        column_num = num_column + 1
        for dic in list_dict:
            for i, element in enumerate(dic.values()):
                if type(element) == float:
                    work.cell(row =i + 2, column = column_num).number_format = '0.00%'
                work.cell(row =i + 2, column = column_num).value = element
                work.cell(row =i + 2, column = column_num).border = style_border
            column_num += 1

    @staticmethod
    def get_width_column(sheet, font_size=11):
        """
        Получение ширины колонок в таблице.
        Args:
            sheet: Лист xlsx-файла
            font_size (int): Шрифт
        """
        columns_dictionary = {}
        for row in sheet.rows:
            for cell in row:
                letter = cell.column_letter
                if cell.value:
                    len_cell = len(str(cell.value))
                    cell_dict = 0
                    if letter in columns_dictionary:
                        cell_dict = columns_dictionary[letter]
                    if len_cell > cell_dict:
                        columns_dictionary[letter] = len_cell
                        new_width_col = len_cell * font_size ** (font_size * 0.009)
                        sheet.column_dimensions[cell.column_letter].width = new_width_col

    def generate_excel_file(self, vacancy_name):
        """
        Создание excel-файла со статистикой по годам и городам.
        Args:
            vacancy_name (str): Название вакансии

        Returns:
            file: excel-файл
        """
        field_year = ['Год',
                      'Средняя зарплата',
                      f'Средняя зарплата - {vacancy_name}',
                      'Количество вакансий',
                      f'Количество вакансий - {vacancy_name}']
        field_city = ['Город',
                      'Уровень зарплат',
                      'Город',
                      'Доля вакансий']
        style = Side(border_style = self.style_border, color = self.color_border)
        bold_heading, border, work_book, work_sheet_city, work_sheet_year = self.prapare_excel_data(
            field_city, field_year, style)
        for x in work_book:
            for cell in x[1]:
                cell.font = bold_heading
                cell.border = border
        work_sheet_city.insert_cols(3, 1)
        Report.get_width_column(work_sheet_year)
        Report.get_width_column(work_sheet_city)

        work_book.save('report.xlsx')

    def prapare_excel_data(self, field_statistic_city, field_statistic_year, style_border):
        """
        Подготовка данных для файла (вспомогательны метод).
        Args:
            field_statistic_city: Поле со статистикой по городам
            field_statistic_year: Поле со статистикой по годам
            style_border (str): Толщина обводки ячеек таблицы
        """
        bold_heading = Font(bold=self.bold_text)
        border = Border(top = style_border, bottom = style_border, left = style_border, right = style_border)
        work_book = Workbook()
        work_year = work_book.active
        work_year.title = self.sheet_title_year
        work_city = work_book.create_sheet(self.sheet_title_city)
        work_year.append(field_statistic_year)
        work_city.append(field_statistic_city)
        Report.append_values(analytic_year, work_year, border)
        Report.append_values(analytic_city, work_city, border)
        Report.append_values(analytic_city_new, work_city, border, 3)
        return bold_heading, border, work_book, work_city, work_year

analytic_year = []
analytic_city = []
analytic_city_new = []

words_dictionary = {
    'name': 'Название',
    'description': 'Описание',
    'key_skills': 'Навыки',
    'experience_id': 'Опыт работы',
    'premium': 'Премиум-вакансия',
    'employer_name': 'Компания',
    'salary_from': 'Нижняя граница вилки оклада',
    'salary_to': 'Верхняя граница вилки оклада',
    'salary_gross': 'Оклад указан до вычета налогов',
    'salary_currency': 'Идентификатор валюты оклада',
    'area_name': 'Название региона',
    'published_at': 'Дата публикации вакансии',
    'salary': 'Оклад'}

work_experience = {
    'noExperience': 'Нет опыта',
    'between1And3': 'От 1 года до 3 лет',
    'between3And6': 'От 3 до 6 лет',
    'moreThan6': 'Более 6 лет'}

new_currency = {'AZN': 'Манаты',
                'BYR': 'Белорусские рубли',
                'EUR': 'Евро',
                'GEL': 'Грузинский лари',
                'KGS': 'Киргизский сом',
                'KZT': 'Тенге', 'RUR': 'Рубли',
                'UAH': 'Гривны',
                'USD': 'Доллары',
                'UZS': 'Узбекский сум'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

experience_order = {
    'noExperience': 0,
    'between1And3': 1,
    'between3And6': 2,
    'moreThan6': 3,
}

bool_to_str = {
    'False': 'Нет',
    'True': 'Да'
}

def prepare_data():
    """
    Подготовка данных для корректой работы программы
    """
    global list_transform_dictionary, list_transform_city, list_transform_city_new
    list_transform_dictionary = []
    list_transform_city = []
    list_transform_city_new = []

prepare_data()

def main():
    """
    Создание объекта InputData, печать данных и графиков.
    """
    a = InputData()
    a.make_data()

if __name__ == '__main__':
    main()




